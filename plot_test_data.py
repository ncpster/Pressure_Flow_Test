# Data Plotting Utility
# Reads Excel files from pressure/flow tests and visualizes the data
# Features:
#   - Browse and load Excel test data files
#   - Display average flow rates from both Alicats
#   - Plot pressure decay from Alicat A over time
#   - Compare multiple test files on the same plot
#   - Clear plot to start fresh comparison

import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook
import os

# Default starting directory for file browser
DEFAULT_DIR = r"C:\Users\patri\RnD\SW Test Data"


class DataPlottingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pressure Flow Test Data Plotter")
        self.root.geometry("1000x700")
        
        # Data storage for multiple files
        self.loaded_files = {}  # {filename: {'avg_flow_a': float, 'avg_flow_b': float, 'time': [], 'pressure': []}}
        self.plot_colors = ['blue', 'red', 'green', 'orange', 'purple', 'brown', 'pink', 'gray']
        self.color_index = 0
        
        # Build GUI
        self.build_gui()
        
    def build_gui(self):
        """Build the GUI interface"""
        # Top control panel
        control_frame = tk.Frame(self.root, bg='lightgray', height=80)
        control_frame.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        control_frame.grid_propagate(False)
        
        tk.Label(control_frame, text="Test Data Plotter", font=('Arial', 16, 'bold'), bg='lightgray').pack(anchor='w', padx=10, pady=5)
        
        # Buttons
        button_subframe = tk.Frame(control_frame, bg='lightgray')
        button_subframe.pack(anchor='w', padx=10, pady=5)
        
        self.load_button = tk.Button(button_subframe, text="Load Excel File", command=self.load_file, 
                                     bg='green', fg='white', width=15, height=1)
        self.load_button.grid(row=0, column=0, padx=5)
        
        self.clear_plot_button = tk.Button(button_subframe, text="Clear Plot", command=self.clear_plot, 
                                          bg='orange', fg='white', width=15, height=1)
        self.clear_plot_button.grid(row=0, column=1, padx=5)
        
        self.remove_last_button = tk.Button(button_subframe, text="Remove Plot", command=self.remove_last_file, 
                                           bg='red', fg='white', width=15, height=1)
        self.remove_last_button.grid(row=0, column=2, padx=5)
        
        # Info panel
        info_frame = tk.LabelFrame(self.root, text="Loaded Files & Average Flow Rates", padx=10, pady=10)
        info_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=10, pady=5)
        
        self.info_text = tk.Text(info_frame, height=8, width=120, state='disabled')
        scrollbar = tk.Scrollbar(info_frame, command=self.info_text.yview)
        self.info_text.config(yscrollcommand=scrollbar.set)
        self.info_text.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Plot area
        plot_frame = tk.LabelFrame(self.root, text="Alicat A Pressure Decay Comparison", padx=10, pady=10)
        plot_frame.grid(row=2, column=0, columnspan=2, sticky='nsew', padx=10, pady=10)
        
        self.fig, self.ax = plt.subplots(figsize=(10, 5))
        self.ax.set_xlabel('Time (s)', fontsize=12)
        self.ax.set_ylabel('Pressure (PSI)', fontsize=12)
        self.ax.set_title('Alicat A Pressure Decay - Multiple Test Comparison')
        self.ax.grid(True, alpha=0.3)
        
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill='both', expand=True)

        self.cursor_label = tk.Label(plot_frame, text="Time: -- s | Pressure: -- PSI", font=('Arial', 12))
        self.cursor_label.pack(anchor='w', padx=5, pady=(5, 0))

        self.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        
        # Configure grid weights
        self.root.grid_rowconfigure(2, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
    def load_file(self):
        """Open file dialog and load Excel file"""
        # Determine starting directory
        start_dir = DEFAULT_DIR if os.path.exists(DEFAULT_DIR) else os.path.expanduser("~")
        
        file_path = filedialog.askopenfilename(
            initialdir=start_dir,
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Select Test Data File"
        )
        
        if not file_path:
            return
        
        try:
            self.parse_excel_file(file_path)
            self.update_plot()
            self.update_info_display()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{e}")
    
    def parse_excel_file(self, file_path):
        """Parse Excel file and extract data"""
        workbook = load_workbook(file_path)
        
        # Get filename for display
        filename = os.path.basename(file_path)
        
        # Check if file already loaded
        if filename in self.loaded_files:
            messagebox.showinfo("Info", f"{filename} is already loaded.")
            return
        
        # Read data from Data sheet
        if "Data" not in workbook.sheetnames:
            raise ValueError("Excel file must contain a 'Data' sheet")
        
        data_sheet = workbook["Data"]
        
        time_data = []
        pressure_data = []
        flow_a_values = []
        flow_b_values = []
        
        # Parse rows (skip header)
        for row_idx, row in enumerate(data_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:  # Skip empty rows
                continue
            
            phase = row[0]
            
            # Only process Pressure Decay phase
            if phase == "Pressure Decay":
                try:
                    time_val = float(row[1]) if row[1] is not None else 0
                    pressure_val = float(row[2]) if row[2] is not None else 0
                    flow_a = float(row[4]) if row[4] is not None else 0
                    flow_b = float(row[5]) if row[5] is not None else 0
                    
                    time_data.append(time_val)
                    pressure_data.append(pressure_val)
                    flow_a_values.append(flow_a)
                    flow_b_values.append(flow_b)
                except (ValueError, TypeError):
                    continue
            elif phase == "Flow Test":
                # Also collect flow data from Flow Test phase
                try:
                    flow_a = float(row[4]) if row[4] is not None else 0
                    flow_b = float(row[5]) if row[5] is not None else 0
                    flow_a_values.append(flow_a)
                    flow_b_values.append(flow_b)
                except (ValueError, TypeError):
                    continue
        
        # Calculate averages
        avg_flow_a = sum(flow_a_values) / len(flow_a_values) if flow_a_values else 0
        avg_flow_b = sum(flow_b_values) / len(flow_b_values) if flow_b_values else 0
        
        # Store data
        self.loaded_files[filename] = {
            'avg_flow_a': avg_flow_a,
            'avg_flow_b': avg_flow_b,
            'time': time_data,
            'pressure': pressure_data,
            'color': self.plot_colors[self.color_index % len(self.plot_colors)]
        }
        
        self.color_index += 1
        
    def update_plot(self):
        """Update the pressure decay plot with all loaded files"""
        self.ax.clear()
        
        if not self.loaded_files:
            self.ax.set_xlabel('Time (s)', fontsize=12)
            self.ax.set_ylabel('Pressure (PSI)', fontsize=12)
            self.ax.set_title('Alicat A Pressure Decay - Multiple Test Comparison')
            self.ax.grid(True, alpha=0.3)
            self.canvas.draw()
            return
        
        # Plot each loaded file
        for filename, data in self.loaded_files.items():
            if data['time'] and data['pressure']:
                self.ax.plot(data['time'], data['pressure'], marker='o', linewidth=2, 
                           label=filename, color=data['color'], markersize=3, alpha=0.7)
        
        self.ax.set_xlabel('Time (s)', fontsize=12)
        self.ax.set_ylabel('Pressure (PSI)', fontsize=12)
        self.ax.set_title('Alicat A Pressure Decay - Multiple Test Comparison')
        self.ax.grid(True, alpha=0.3)
        self.ax.legend(loc='best', fontsize=10)
        
        # Set reasonable y-axis limits
        all_pressures = []
        for data in self.loaded_files.values():
            all_pressures.extend(data['pressure'])
        if all_pressures:
            min_p = min(all_pressures)
            max_p = max(all_pressures)
            padding = (max_p - min_p) * 0.1 or 1
            self.ax.set_ylim(min_p - padding, max_p + padding)
        
        self.canvas.draw()

    def on_mouse_move(self, event):
        """Update cursor readout with time and pressure under the mouse"""
        if event.inaxes != self.ax or event.xdata is None or event.ydata is None:
            self.cursor_label.config(text="Time: -- s | Pressure: -- PSI")
            return

        self.cursor_label.config(text=f"Time: {event.xdata:.2f} s | Pressure: {event.ydata:.2f} PSI")
    
    def update_info_display(self):
        """Update the info text display with loaded files and flow data"""
        self.info_text.config(state='normal')
        self.info_text.delete('1.0', tk.END)
        
        # Configure color tags for each file
        for filename, data in self.loaded_files.items():
            color = data['color']
            self.info_text.tag_config(f"color_{filename}", foreground=color, font=('Courier', 12, 'bold'))
        
        if not self.loaded_files:
            self.info_text.insert(tk.END, "No files loaded. Click 'Load Excel File' to get started.")
        else:
            # Calculate column widths based on actual data
            max_filename_len = max(len("File Name"), max(len(f) for f in self.loaded_files.keys()))
            col1_width = max_filename_len + 2
            col2_width = 20
            col3_width = 20
            
            # Create header
            header = f"{'File Name':<{col1_width}} {'Avg Flow A (SLPM)':<{col2_width}} {'Avg Flow B (SLPM)':<{col3_width}}\n"
            header += "-" * (col1_width + col2_width + col3_width) + "\n"
            self.info_text.insert(tk.END, header)
            
            for filename, data in self.loaded_files.items():
                # Insert filename with color tag
                filename_str = f"{filename:<{col1_width}}"
                self.info_text.insert(tk.END, filename_str, f"color_{filename}")
                
                # Insert flow data in normal text
                flow_str = f"{data['avg_flow_a']:<{col2_width}.3f} {data['avg_flow_b']:<{col3_width}.3f}\n"
                self.info_text.insert(tk.END, flow_str)
        
        self.info_text.config(state='disabled')
    
    def clear_plot(self):
        """Clear all loaded files and reset plot"""
        if not self.loaded_files:
            messagebox.showinfo("Info", "No files loaded to clear.")
            return
        
        self.loaded_files.clear()
        self.color_index = 0
        self.update_plot()
        self.update_info_display()
        messagebox.showinfo("Success", "Plot cleared. Ready to load new files.")
    
    def remove_last_file(self):
        """Show a dialog to select which file to remove"""
        if not self.loaded_files:
            messagebox.showinfo("Info", "No files loaded.")
            return
        
        # Create a selection dialog
        file_list = list(self.loaded_files.keys())
        
        # Create a simple Toplevel window for selection
        selection_window = tk.Toplevel(self.root)
        selection_window.title("Select File to Remove")
        selection_window.geometry("500x300")
        selection_window.transient(self.root)
        selection_window.grab_set()
        
        # Label
        tk.Label(selection_window, text="Select a file to remove:", font=('Arial', 12)).pack(padx=10, pady=10)
        
        # Listbox with scrollbar
        frame = tk.Frame(selection_window)
        frame.pack(padx=10, pady=10, fill='both', expand=True)
        
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, font=('Courier', 11))
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        
        # Add files to listbox with their colors
        for filename in file_list:
            listbox.insert(tk.END, filename)
            color = self.loaded_files[filename]['color']
            listbox.itemconfig(tk.END, fg=color)
        
        # Select first item by default
        if listbox.size() > 0:
            listbox.selection_set(0)
            listbox.see(0)
        
        # Buttons
        button_frame = tk.Frame(selection_window)
        button_frame.pack(padx=10, pady=10)
        
        def remove_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Warning", "Please select a file to remove.")
                return
            
            selected_file = file_list[selection[0]]
            del self.loaded_files[selected_file]
            self.update_plot()
            self.update_info_display()
            selection_window.destroy()
        
        def cancel():
            selection_window.destroy()
        
        tk.Button(button_frame, text="Remove", command=remove_selected, bg='red', fg='white', width=12, padx=10).pack(side='left', padx=5)
        tk.Button(button_frame, text="Cancel", command=cancel, bg='gray', fg='white', width=12, padx=10).pack(side='left', padx=5)


# Create and run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = DataPlottingApp(root)
    root.mainloop()
